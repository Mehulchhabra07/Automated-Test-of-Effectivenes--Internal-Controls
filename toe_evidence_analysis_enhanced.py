"""
------------------------------------------------------------------
AI-Powered Test of Effectiveness (TOE) Evidence Analysis Tool
requires : pandas  openpyxl  openai  httpx  python-docx  PyPDF2  extract-msg  requests  pytesseract  Pillow  pdf2image
------------------------------------------------------------------
An intelligent auditing framework that uses AI to analyze evidence files 
for effectiveness testing. Features automated evidence processing, 
AI-driven analysis, and comprehensive reporting with Excel output.

This tool automates the traditionally manual process of evidence evaluation
for Test of Effectiveness procedures, reducing analysis time from hours to 
minutes while maintaining high accuracy and consistency.

Author: Mehul Chhabra
GitHub: https://github.com/Mehulchhabra07/AI-Powered-TOE-Evidence-Analysis
Project: AI-Driven Evidence Testing Framework
------------------------------------------------------------------
"""

from pathlib import Path
from datetime import datetime
import os, sys, json, time, re, logging
import pandas as pd
import httpx
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openai import OpenAI, OpenAIError
from typing import Tuple, Optional, Dict, Any, List
try:
    from docx import Document
except ImportError:
    print("Warning: python-docx not installed. DOCX files will be skipped.")
    Document = None
try:
    import PyPDF2
except ImportError:
    print("Warning: PyPDF2 not installed. PDF files will be skipped.")
    PyPDF2 = None
try:
    import openpyxl
except ImportError:
    print("Warning: openpyxl not installed for reading Excel files as evidence.")

try:
    import extract_msg
except ImportError:
    print("Warning: extract-msg not installed. MSG files will be skipped.")
    print("Install with: pip install extract-msg")
    extract_msg = None

try:
    import email
    from email import policy
    from email.parser import BytesParser
except ImportError:
    print("Warning: email module not available. EML files will be skipped.")
    email = None

try:
    import pytesseract
    from PIL import Image
    import pdf2image
except ImportError:
    print("Warning: OCR libraries not installed. Image files and image-based PDFs will have limited text extraction.")
    print("Install with: pip install pytesseract pillow pdf2image")
    print("Also install Tesseract OCR: https://github.com/UB-Mannheim/tesseract/wiki")
    pytesseract = None
    Image = None
    pdf2image = None

# ═══════════════════════════════════════════════════════════════════════════════
#                               CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

class Config:
    """Configuration settings for TOE Evidence Analysis"""
    
    # File locations - Update these paths for your environment
    INPUT_FILE = Path("sample_controls.xlsx")  # Place your input file in the same directory
    OUTPUT_FILE = INPUT_FILE.with_name(f"{INPUT_FILE.stem}_TOE_EvidenceAnalysis.xlsx")
    
    # Evidence folder path
    EVIDENCE_ROOT = "Evidence"  # Folder containing evidence subfolders for each control
    
    # Input columns (exact headers in row‑1)
    CONTROL_COL = "Control"
    DESC_COL = "Control Description"
    
    # OpenAI API settings - Set your API key here or via environment variable
    API_KEY = os.getenv("OPENAI_API_KEY", "YOUR_OPENAI_API_KEY_HERE")
    BASE_URL = "https://api.openai.com/v1"  # Standard OpenAI API endpoint
    MODEL = "gpt-4o"  # Using GPT-4o for better analysis quality
    
    # Supported OpenAI models
    SUPPORTED_MODELS = [
        "gpt-4", "gpt-4-turbo", "gpt-3.5-turbo", 
        "gpt-4o", "gpt-4o-mini"
    ]
    
    # Retry settings for robust error handling
    MAX_RETRIES = 5
    RETRY_DELAY = 1.0
    MAX_RETRY_DELAY = 60.0
    
    # Request timeout settings
    REQUEST_TIMEOUT = 120.0
    
    # Required columns for validation
    REQUIRED_COLS = [
        "Risk", "Risk Description", "Control", "Control Description"
    ]
    
    # SAP GRC Integration Settings
    SAP_GRC_ENABLED = False  # Set to True to enable SAP GRC integration
    SAP_GRC_URL = ""  # Your SAP GRC system URL
    SAP_GRC_USERNAME = os.getenv("SAP_GRC_USER", "")
    SAP_GRC_PASSWORD = os.getenv("SAP_GRC_PASS", "")
    
    # Jira Integration Settings
    JIRA_ENABLED = False  # Set to True to enable Jira integration
    JIRA_URL = ""  # Your Jira instance URL (e.g., https://yourcompany.atlassian.net)
    JIRA_USERNAME = os.getenv("JIRA_USER", "")
    JIRA_TOKEN = os.getenv("JIRA_TOKEN", "")

# ═══════════════════════════════════════════════════════════════════════════════
#                               LOGGING SETUP
# ═══════════════════════════════════════════════════════════════════════════════

def setup_logging():
    """Setup logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('toe_evidence_analysis.log'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ═══════════════════════════════════════════════════════════════════════════════
#                               API CLIENT SETUP
# ═══════════════════════════════════════════════════════════════════════════════

def initialize_client() -> OpenAI:
    """Initialize and test OpenAI client with robust error handling"""
    if Config.API_KEY == "YOUR_OPENAI_API_KEY_HERE":
        logger.error("✗ OpenAI API key not configured. Set OPENAI_API_KEY environment variable.")
        sys.exit(1)
    
    client = OpenAI(api_key=Config.API_KEY, base_url=Config.BASE_URL)
    
    # Test the connection
    for attempt in range(Config.MAX_RETRIES):
        try:
            logger.info(f"Testing connection (attempt {attempt + 1}/{Config.MAX_RETRIES})...")
            response = client.chat.completions.create(
                model=Config.MODEL,
                messages=[
                    {"role": "system", "content": "Be concise and precise."},
                    {"role": "user", "content": "ping"}
                ],
                max_tokens=10
            )
            
            if response and response.choices and response.choices[0].message.content:
                logger.info("✔ OpenAI API connection test successful")
                logger.info(f"✔ Model '{Config.MODEL}' is working correctly")
                return client
            else:
                raise Exception("Invalid response structure from API")
                
        except Exception as e:
            error_msg = str(e).lower()
            logger.warning(f"Connection test attempt {attempt + 1} failed: {e}")
            
            # Handle specific error types
            if "unauthorized" in error_msg or "invalid api key" in error_msg:
                logger.error("✗ Invalid API key. Please check your OPENAI_API_KEY.")
                sys.exit(1)
            elif "model" in error_msg and "not found" in error_msg:
                logger.error(f"✗ Model '{Config.MODEL}' not available. Check supported models.")
                sys.exit(1)
            
            if attempt < Config.MAX_RETRIES - 1:
                wait_time = Config.RETRY_DELAY * (2 ** attempt)
                logger.info(f"Retrying in {wait_time:.1f} seconds...")
                time.sleep(wait_time)
            else:
                logger.error("✗ Failed to connect to OpenAI API after all retries")
                sys.exit(1)
    
    return client

# ═══════════════════════════════════════════════════════════════════════════════
#                               UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

# Additional imports for SAP GRC and Jira integration
try:
    import requests
    from requests.auth import HTTPBasicAuth
except ImportError:
    print("Warning: requests not installed. SAP GRC and Jira integration will be disabled.")
    print("Install with: pip install requests")
    requests = None

# Token Management Settings
MAX_FILE_CHARS = 15000         # Max chars per individual file (allows full analysis)
MAX_TOTAL_CHARS = 160000       # Max total chars for all evidence combined (~40k tokens)

BRACE_RE = re.compile(r"\{.*\}", re.S)

def extract_json_from_response(response: str) -> Optional[Dict[str, Any]]:
    """Extract JSON from LLM response"""
    try:
        match = BRACE_RE.search(response)
        if match:
            return json.loads(match.group())
        return json.loads(response)
    except:
        return None

def make_llm_call_with_retry(client: OpenAI, prompt: str) -> Optional[str]:
    """Make LLM call with retry logic"""
    for attempt in range(Config.MAX_RETRIES):
        try:
            response = client.chat.completions.create(
                model=Config.MODEL,
                messages=[
                    {"role": "system", "content": "You are an expert auditor analyzing evidence. Provide precise, professional analysis."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.3,
                timeout=Config.REQUEST_TIMEOUT
            )
            
            if response and response.choices and response.choices[0].message.content:
                return response.choices[0].message.content.strip()
            else:
                logger.warning(f"Empty response from API (attempt {attempt + 1})")
                
        except OpenAIError as e:
            logger.warning(f"OpenAI API error (attempt {attempt + 1}): {e}")
            if attempt < Config.MAX_RETRIES - 1:
                wait_time = min(Config.RETRY_DELAY * (2 ** attempt), Config.MAX_RETRY_DELAY)
                logger.info(f"Retrying in {wait_time:.1f} seconds...")
                time.sleep(wait_time)
            else:
                logger.error("Max retries exceeded for LLM call")
                
        except Exception as e:
            logger.error(f"Unexpected error in LLM call: {e}")
            break
    
    return None

# ═══════════════════════════════════════════════════════════════════════════════
#                           INTEGRATION FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_sap_grc_evidence(control_id: str) -> str:
    """Fetch evidence from SAP GRC system for a specific control"""
    if not Config.SAP_GRC_ENABLED:
        return ""
    
    try:
        import requests
        from requests.auth import HTTPBasicAuth
        
        # This is a template - adjust based on your SAP GRC REST API endpoints
        auth = HTTPBasicAuth(Config.SAP_GRC_USERNAME, Config.SAP_GRC_PASSWORD)
        
        # Example: Fetch control documentation
        response = requests.get(
            f"{Config.SAP_GRC_URL}/sap/bc/rest/grc/controls/{control_id}",
            auth=auth,
            headers={'Accept': 'application/json'},
            timeout=30
        )
        
        if response.status_code == 200:
            data = response.json()
            evidence_text = []
            evidence_text.append("=== SAP GRC CONTROL DATA ===")
            evidence_text.append(f"Control ID: {data.get('control_id', 'N/A')}")
            evidence_text.append(f"Control Name: {data.get('control_name', 'N/A')}")
            evidence_text.append(f"Status: {data.get('status', 'N/A')}")
            evidence_text.append(f"Last Review Date: {data.get('last_review_date', 'N/A')}")
            evidence_text.append(f"Reviewer: {data.get('reviewer', 'N/A')}")
            evidence_text.append(f"Control Description: {data.get('description', 'N/A')}")
            
            # Fetch test results if available
            test_response = requests.get(
                f"{Config.SAP_GRC_URL}/sap/bc/rest/grc/controls/{control_id}/tests",
                auth=auth,
                headers={'Accept': 'application/json'},
                timeout=30
            )
            
            if test_response.status_code == 200:
                test_data = test_response.json()
                evidence_text.append("\n=== CONTROL TEST RESULTS ===")
                for test in test_data.get('tests', []):
                    evidence_text.append(f"Test Date: {test.get('test_date', 'N/A')}")
                    evidence_text.append(f"Test Result: {test.get('result', 'N/A')}")
                    evidence_text.append(f"Tester: {test.get('tester', 'N/A')}")
                    evidence_text.append(f"Comments: {test.get('comments', 'N/A')}")
                    evidence_text.append("-" * 30)
            
            return '\n'.join(evidence_text)
        else:
            return f"[SAP GRC Error: HTTP {response.status_code}]"
            
    except Exception as e:
        return f"[SAP GRC Error: {e}]"

def get_jira_evidence(control_name: str) -> str:
    """Fetch evidence from Jira tickets related to a control"""
    if not Config.JIRA_ENABLED:
        return ""
    
    try:
        import requests
        from requests.auth import HTTPBasicAuth
        
        auth = HTTPBasicAuth(Config.JIRA_USERNAME, Config.JIRA_TOKEN)
        
        # Search for Jira tickets related to the control
        search_query = f'summary~"{control_name}" OR description~"{control_name}"'
        
        response = requests.get(
            f"{Config.JIRA_URL}/rest/api/3/search",
            auth=auth,
            headers={'Accept': 'application/json'},
            params={
                'jql': search_query,
                'maxResults': 10,
                'fields': 'summary,description,status,assignee,created,updated,comments'
            },
            timeout=30
        )
        
        if response.status_code == 200:
            data = response.json()
            evidence_text = []
            evidence_text.append("=== JIRA TICKET EVIDENCE ===")
            
            for issue in data.get('issues', []):
                fields = issue.get('fields', {})
                evidence_text.append(f"\nTicket: {issue.get('key', 'N/A')}")
                evidence_text.append(f"Summary: {fields.get('summary', 'N/A')}")
                evidence_text.append(f"Status: {fields.get('status', {}).get('name', 'N/A')}")
                evidence_text.append(f"Assignee: {fields.get('assignee', {}).get('displayName', 'Unassigned') if fields.get('assignee') else 'Unassigned'}")
                evidence_text.append(f"Created: {fields.get('created', 'N/A')}")
                evidence_text.append(f"Updated: {fields.get('updated', 'N/A')}")
                
                description = fields.get('description', '')
                if description:
                    # Handle description content (might be in ADF format)
                    if isinstance(description, dict):
                        desc_text = extract_text_from_adf(description)
                    else:
                        desc_text = str(description)
                    
                    if len(desc_text) > 500:
                        desc_text = desc_text[:500] + "[TRUNCATED]"
                    evidence_text.append(f"Description: {desc_text}")
                
                # Add comments if available
                comments = fields.get('comment', {}).get('comments', [])
                if comments:
                    evidence_text.append("Recent Comments:")
                    for comment in comments[-3:]:  # Last 3 comments
                        author = comment.get('author', {}).get('displayName', 'Unknown')
                        created = comment.get('created', 'Unknown')
                        body = comment.get('body', '')
                        if isinstance(body, dict):
                            body = extract_text_from_adf(body)
                        if len(body) > 200:
                            body = body[:200] + "[TRUNCATED]"
                        evidence_text.append(f"  - {author} ({created}): {body}")
                
                evidence_text.append("-" * 40)
            
            if not data.get('issues'):
                evidence_text.append("No related Jira tickets found.")
            
            return '\n'.join(evidence_text)
        else:
            return f"[Jira Error: HTTP {response.status_code}]"
            
    except Exception as e:
        return f"[Jira Error: {e}]"

def extract_text_from_adf(adf_content):
    """Extract text from Atlassian Document Format (ADF)"""
    if not isinstance(adf_content, dict):
        return str(adf_content)
    
    text_parts = []
    
    def extract_text_recursive(node):
        if isinstance(node, dict):
            if node.get('type') == 'text':
                text_parts.append(node.get('text', ''))
            elif 'content' in node:
                for child in node['content']:
                    extract_text_recursive(child)
        elif isinstance(node, list):
            for item in node:
                extract_text_recursive(item)
    
    extract_text_recursive(adf_content)
    return ' '.join(text_parts)

# ───── Evidence File Reading Utilities ────────────────────────
def read_txt_file(path):
    """Read plain text files"""
    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()

def read_docx_file(path):
    """Read Microsoft Word documents"""
    if Document is None:
        return "[DOCX file - python-docx not installed]"
    try:
        doc = Document(path)
        return '\n'.join([p.text for p in doc.paragraphs])
    except Exception as e:
        return f"[Error reading DOCX: {e}]"

def read_image_file(path):
    """Read image files using OCR with enhanced content management"""
    if pytesseract is None or Image is None:
        return "[OCR not available - install pytesseract, pillow, and Tesseract OCR]"
    try:
        image = Image.open(path)
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(image, config=custom_config)
        data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT, config=custom_config)
        confidences = [int(conf) for conf in data['conf'] if int(conf) > 0]
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0
        
        # Allow more OCR text for better analysis
        if len(text) > 8000:
            text = text[:7500] + f"\n[OCR TRUNCATED: Full text was {len(text)} chars]"
        
        result = f"OCR Text (Confidence: {avg_confidence:.1f}%):\n{text}"
        if avg_confidence < 70:
            result += f"\n[WARNING: Low OCR confidence - text may be inaccurate]"
        return result
    except Exception as e:
        return f"[Error performing OCR on image: {e}]"

def extract_text_from_pdf_with_ocr(pdf_path):
    """Extract text from PDF using both PyPDF2 and OCR as fallback with enhanced token management"""
    text_content = []
    if PyPDF2 is not None:
        try:
            with open(pdf_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                extracted_text = ""
                max_pages = min(10, len(reader.pages))  # Increased to 10 pages for better analysis
                
                for page_num in range(max_pages):
                    page = reader.pages[page_num]
                    page_text = page.extract_text() or ''
                    extracted_text += page_text
                    # Stop if we have substantial content
                    if len(extracted_text) > 12000:
                        extracted_text = extracted_text[:12000]
                        break
                
                if extracted_text.strip():
                    text_content.append(f"=== PDF TEXT EXTRACTION ({max_pages} pages) ===")
                    text_content.append(extracted_text)
                    if len(reader.pages) > max_pages:
                        text_content.append(f"[NOTE: PDF has {len(reader.pages)} total pages, processed first {max_pages}]")
                    return '\n'.join(text_content)
                else:
                    text_content.append("[PDF contains no extractable text - attempting OCR...]")
        except Exception as e:
            text_content.append(f"[PDF text extraction failed: {e} - attempting OCR...]")
    
    if pytesseract is not None and pdf2image is not None:
        try:
            text_content.append("\n=== PDF OCR EXTRACTION ===")
            # Increase to first 8 pages for OCR for better coverage
            images = pdf2image.convert_from_path(pdf_path, dpi=200, first_page=1, last_page=8)
            
            for page_num, image in enumerate(images):
                text_content.append(f"\n--- PAGE {page_num + 1} OCR ---")
                custom_config = r'--oem 3 --psm 6'
                page_text = pytesseract.image_to_string(image, config=custom_config)
                
                # Allow more OCR text per page for better analysis
                if len(page_text) > 2000:
                    page_text = page_text[:1800] + f"\n[PAGE {page_num + 1} OCR TRUNCATED]"
                
                data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT, config=custom_config)
                confidences = [int(conf) for conf in data['conf'] if int(conf) > 0]
                avg_confidence = sum(confidences) / len(confidences) if confidences else 0
                text_content.append(f"OCR Confidence: {avg_confidence:.1f}%")
                text_content.append(page_text)
                
                if avg_confidence < 70:
                    text_content.append(f"[WARNING: Low OCR confidence for this page]")
            
            if len(images) == 8:
                text_content.append(f"[NOTE: OCR processed first 8 pages for comprehensive analysis]")
            
            return '\n'.join(text_content)
        except Exception as e:
            text_content.append(f"[PDF OCR failed: {e}]")
            return '\n'.join(text_content)
    else:
        text_content.append("[OCR not available - install OCR libraries for image-based PDF processing]")
        return '\n'.join(text_content)

def read_pdf_file(path):
    """Read PDF files with OCR fallback"""
    return extract_text_from_pdf_with_ocr(path)

def read_csv_file(path):
    """Read CSV files"""
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    except Exception as e:
        return f"[Error reading CSV: {e}]"

def read_xlsx_file(path):
    """Read Excel files"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        text = []
        for ws in wb.worksheets:
            text.append(f"--- Sheet: {ws.title} ---")
            for row in ws.iter_rows(values_only=True):
                text.append(', '.join([str(cell) for cell in row if cell is not None]))
        return '\n'.join(text)
    except Exception as e:
        return f"[Error reading XLSX: {e}]"

def read_msg_file(path):
    """Read Outlook MSG files with content summarization"""
    if extract_msg is None:
        return "[MSG file - extract-msg not installed. Install with: pip install extract-msg]"
    try:
        msg = extract_msg.Message(path)
        
        # Extract email content
        content = []
        content.append(f"From: {msg.sender}")
        content.append(f"To: {msg.to}")
        content.append(f"CC: {msg.cc or 'None'}")
        content.append(f"Date: {msg.date}")
        content.append(f"Subject: {msg.subject}")
        content.append("-" * 50)
        content.append("Body:")
        
        # Limit email body content to prevent token overuse
        body = msg.body or "[No body content]"
        if len(body) > 800:
            body = body[:700] + f"\n[EMAIL BODY TRUNCATED - original length: {len(body)} chars]"
        content.append(body)
        
        # Add attachment info if any
        if msg.attachments:
            content.append("\n" + "-" * 50)
            content.append("Attachments:")
            for i, attachment in enumerate(msg.attachments[:5]):  # Limit to first 5 attachments
                content.append(f"  - {attachment.longFilename or attachment.shortFilename}")
            if len(msg.attachments) > 5:
                content.append(f"  ... and {len(msg.attachments) - 5} more attachments")
        
        return '\n'.join(content)
    except Exception as e:
        return f"[Error reading MSG: {e}]"

def read_eml_file(path):
    """Read EML email files"""
    if email is None:
        return "[EML file - email module not available]"
    try:
        with open(path, 'rb') as f:
            msg = BytesParser(policy=policy.default).parse(f)
        
        content = []
        content.append(f"From: {msg.get('From', 'Unknown')}")
        content.append(f"To: {msg.get('To', 'Unknown')}")
        content.append(f"CC: {msg.get('CC', 'None')}")
        content.append(f"Date: {msg.get('Date', 'Unknown')}")
        content.append(f"Subject: {msg.get('Subject', 'No Subject')}")
        content.append("-" * 50)
        content.append("Body:")
        
        # Get email body
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    content.append(part.get_content())
                    break
                elif part.get_content_type() == "text/html":
                    # Basic HTML stripping for readability
                    html_content = part.get_content()
                    import re
                    clean_content = re.sub('<[^<]+?>', '', html_content)
                    content.append(clean_content)
                    break
        else:
            content.append(msg.get_content())
        
        return '\n'.join(content)
    except Exception as e:
        return f"[Error reading EML: {e}]"

def read_mbox_file(path):
    """Read MBOX email files"""
    try:
        import mailbox
        content = []
        mbox = mailbox.mbox(path)
        
        for i, message in enumerate(mbox):
            content.append(f"\n=== EMAIL {i+1} ===")
            content.append(f"From: {message.get('From', 'Unknown')}")
            content.append(f"To: {message.get('To', 'Unknown')}")
            content.append(f"Date: {message.get('Date', 'Unknown')}")
            content.append(f"Subject: {message.get('Subject', 'No Subject')}")
            content.append("-" * 30)
            
            # Get email body
            if message.is_multipart():
                for part in message.walk():
                    if part.get_content_type() == "text/plain":
                        content.append(part.get_payload(decode=True).decode('utf-8', errors='ignore'))
                        break
            else:
                content.append(message.get_payload(decode=True).decode('utf-8', errors='ignore'))
        
        return '\n'.join(content)
    except Exception as e:
        return f"[Error reading MBOX: {e}]"

def estimate_tokens(text: str) -> int:
    """Rough estimation of tokens (approximately 4 characters = 1 token)"""
    return len(text) // 4

def smart_truncate_content(content: str, max_chars: int) -> str:
    """Smart truncation that preserves key information"""
    if len(content) <= max_chars:
        return content
    
    # Try to keep the beginning and end, removing middle
    keep_start = max_chars // 2
    keep_end = max_chars // 4
    
    if keep_start + keep_end < max_chars:
        truncated = (content[:keep_start] + 
                    f"\n\n[TRUNCATED: {len(content) - max_chars} chars removed from middle]\n\n" + 
                    content[-keep_end:])
        return truncated[:max_chars]
    else:
        return content[:max_chars] + f"\n[TRUNCATED: content exceeds {max_chars} chars]"

def process_large_excel_file(path: str) -> str:
    """Process Excel files with smart summarization for large files"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        
        summary = []
        summary.append(f"Excel Workbook: {len(wb.worksheets)} sheets")
        
        for ws in wb.worksheets:
            sheet_info = [f"--- Sheet: {ws.title} ---"]
            row_count = 0
            
            # Read first 10 rows and last 5 rows for each sheet
            rows_data = []
            for row in ws.iter_rows(values_only=True, max_row=15):
                if row_count < 10 or row_count >= ws.max_row - 5:
                    if any(cell is not None for cell in row):
                        rows_data.append(', '.join([str(cell) for cell in row if cell is not None]))
                row_count += 1
                if row_count > 15:  # Limit processing
                    break
            
            if row_count > 15:
                rows_data.append(f"[SUMMARY: Sheet has {ws.max_row} total rows, showing first 10 and last 5]")
            
            sheet_info.extend(rows_data)
            summary.append('\n'.join(sheet_info))
        
        return '\n'.join(summary)
    except Exception as e:
        return f"[Error processing Excel file: {e}]"

def read_evidence_folder(control_name: str) -> str:
    """Read all evidence files from a control's folder with SAP GRC and Jira integration."""
    clean_name = re.sub(r'[<>:"/\\|?*]', '_', str(control_name).strip())
    folder = os.path.join(Config.EVIDENCE_ROOT, clean_name)
    
    # Try to find the folder with fuzzy matching
    if not os.path.isdir(folder):
        for fname in os.listdir(Config.EVIDENCE_ROOT) if os.path.exists(Config.EVIDENCE_ROOT) else []:
            if os.path.isdir(os.path.join(Config.EVIDENCE_ROOT, fname)) and clean_name.lower() in fname.lower():
                folder = os.path.join(Config.EVIDENCE_ROOT, fname)
                break
        else:
            # Create basic folder structure info if no evidence folder found
            folder_missing = True
    else:
        folder_missing = False
    
    contents = []
    file_count = 0
    total_chars = 0
    truncated = False
    
    # Add SAP GRC evidence if enabled
    if Config.SAP_GRC_ENABLED:
        print("  🔍 Fetching SAP GRC evidence...")
        sap_evidence = get_sap_grc_evidence(control_name)
        if sap_evidence:
            contents.append(sap_evidence)
            total_chars += len(sap_evidence)
            file_count += 1
    
    # Add Jira evidence if enabled
    if Config.JIRA_ENABLED:
        print("  🎫 Fetching Jira evidence...")
        jira_evidence = get_jira_evidence(control_name)
        if jira_evidence:
            contents.append(jira_evidence)
            total_chars += len(jira_evidence)
            file_count += 1
    
    # Process local evidence files if folder exists
    if not folder_missing and os.path.exists(folder):
        # Get all files and sort by size (smallest first to ensure we get more files)
        files = []
        for fname in os.listdir(folder):
            fpath = os.path.join(folder, fname)
            if os.path.isfile(fpath):
                files.append((fname, fpath, os.path.getsize(fpath)))
        
        files.sort(key=lambda x: x[2])  # Sort by file size
        
        for fname, fpath, file_size in files:
            if total_chars >= MAX_TOTAL_CHARS:
                break
                
            file_count += 1
            ext = os.path.splitext(fname)[1].lower()
            contents.append(f"=== LOCAL FILE {file_count}: {fname} ({file_size} bytes) ===")
            
            try:
                if ext == '.txt':
                    file_content = read_txt_file(fpath)
                elif ext == '.docx':
                    file_content = read_docx_file(fpath)
                elif ext == '.pdf':
                    file_content = read_pdf_file(fpath)
                elif ext == '.csv':
                    file_content = read_csv_file(fpath)
                elif ext in ['.xlsx', '.xls']:
                    # Use smart Excel processing for large files
                    if file_size > 50000:  # 50KB threshold
                        file_content = process_large_excel_file(fpath)
                    else:
                        file_content = read_xlsx_file(fpath)
                elif ext == '.msg':
                    file_content = read_msg_file(fpath)
                elif ext == '.eml':
                    file_content = read_eml_file(fpath)
                elif ext == '.mbox':
                    file_content = read_mbox_file(fpath)
                elif ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif']:
                    file_content = read_image_file(fpath)
                else:
                    file_content = f"[Unsupported file type: {ext}]"
            except Exception as e:
                file_content = f"[Error reading {fname}: {e}]"
            
            # Smart truncation for individual files
            if len(file_content) > MAX_FILE_CHARS:
                file_content = smart_truncate_content(file_content, MAX_FILE_CHARS)
                truncated = True
            
            # Check total limit
            remaining_chars = MAX_TOTAL_CHARS - total_chars
            if len(file_content) > remaining_chars:
                if remaining_chars > 100:  # Only add if we have meaningful space left
                    file_content = smart_truncate_content(file_content, remaining_chars - 50)
                    contents.append(file_content)
                    contents.append(f"[STOPPED: Total evidence limit of {MAX_TOTAL_CHARS} chars reached]")
                truncated = True
                break
            
            contents.append(file_content)
            total_chars += len(file_content)
            contents.append("")  # Add separator between files
    
    if file_count == 0:
        return f"No evidence found for control: {control_name} (Local folder: {folder}, SAP GRC: {'Enabled' if Config.SAP_GRC_ENABLED else 'Disabled'}, Jira: {'Enabled' if Config.JIRA_ENABLED else 'Disabled'})"
    
    result = '\n'.join(contents)
    
    # Final safety check
    if len(result) > MAX_TOTAL_CHARS:
        result = smart_truncate_content(result, MAX_TOTAL_CHARS)
        truncated = True
    
    if truncated:
        warning = f"[WARNING: Evidence content was truncated. Processed {file_count} sources, total {len(result)} chars]"
        result = warning + "\n\n" + result
    
    print(f"  📄 Processed {file_count} evidence sources, {len(result)} characters")
    print(f"  🔢 Estimated tokens: ~{estimate_tokens(result)}")
    
    return result

# ═══════════════════════════════════════════════════════════════════════════════
#                           LLM EVIDENCE ANALYSIS FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def make_llm_request_with_retry(client: OpenAI, payload: dict, operation_name: str) -> str:
    """Make LLM request with robust retry logic and error handling using OpenAI client"""
    for attempt in range(Config.MAX_RETRIES):
        try:
            # Pre-flight token check
            total_content = ""
            for msg in payload.get("messages", []):
                total_content += msg.get("content", "")
            
            estimated_tokens = estimate_tokens(total_content)
            if estimated_tokens > 15000:  # Conservative limit
                print(f"  ⚠️ Warning: Estimated {estimated_tokens} tokens for {operation_name}")
            
            # Use OpenAI client for API call
            completion = client.chat.completions.create(
                model=payload["model"],
                messages=payload["messages"],
                max_tokens=payload.get("max_tokens", 2048),
                temperature=payload.get("temperature", 0.7)
            )
            
            # Safely extract content
            content = completion.choices[0].message.content
            return content.strip() if content else "No response generated"
                
        except Exception as e:
            error_msg = str(e).lower()
            print(f"  ⚠️ {operation_name} attempt {attempt + 1}/{Config.MAX_RETRIES} failed: {e}")
            
            # Handle specific error types
            if "429" in error_msg or "too many requests" in error_msg or "rate limit" in error_msg:
                wait_time = min(Config.RETRY_DELAY * (2 ** attempt), Config.MAX_RETRY_DELAY)
                print(f"  ⏳ Rate limit hit, waiting {wait_time}s before retry...")
                time.sleep(wait_time)
            elif "401" in error_msg or "unauthorized" in error_msg:
                print(f"  🚫 Authentication failed for {operation_name}")
                return f"Authentication error: {e}"
            elif "400" in error_msg or "bad request" in error_msg:
                print(f"  🚫 Bad request for {operation_name} - possibly too many tokens")
                return f"Request too large: {e}"
            elif attempt < Config.MAX_RETRIES - 1:
                wait_time = Config.RETRY_DELAY * (2 ** attempt)
                print(f"  ⏳ Retrying {operation_name} in {wait_time}s...")
                time.sleep(wait_time)
            else:
                print(f"  🚫 All {Config.MAX_RETRIES} attempts failed for {operation_name}")
    
    return f"LLM error after {Config.MAX_RETRIES} attempts"

def ask_evidence_summary(client: OpenAI, evidence_content: str) -> str:
    """Analyze evidence and provide detailed summary using OpenAI API"""
    # Final safety check on content size
    if len(evidence_content) > MAX_TOTAL_CHARS:
        evidence_content = smart_truncate_content(evidence_content, MAX_TOTAL_CHARS)
        print(f"  ⚠️ Evidence content further truncated to {len(evidence_content)} chars")
    
    prompt = f"""
You are a world-class auditor with decades of experience. Carefully review the following evidence files and provide a comprehensive, detailed bullet-point summary of ALL key information, data, and documentation present.

Be thorough and specific. Structure your analysis as follows:

DOCUMENTS PRESENT:
• [List a general high level overview of files found]

KEY INFORMATION IDENTIFIED:
• [Important data points, dates, amounts, names, IDs]
• [Process steps documented]
• [Approvals, signatures, timestamps]
• [Controls and checkpoints evident]
• [Completeness and accuracy of documentation]

GAPS OR CONCERNS:
• [Any missing information or concerns noted]

EVIDENCE FILES:
{evidence_content}

Provide your structured bullet-point analysis:
"""
    
    payload = {
        "model": Config.MODEL,
        "messages": [
            {"role": "system", "content": "You are an expert auditor. Provide a comprehensive structured analysis of evidence with clear sections."},
            {"role": "user", "content": prompt}
        ],
        "max_tokens": 2048,
        "temperature": 0.7
    }
    
    return make_llm_request_with_retry(client, payload, "Evidence Summary")

def ask_evidence_sufficiency(client: OpenAI, control_desc: str, evidence_content: str) -> str:
    """Assess if evidence is sufficient to pass the control using OpenAI API"""
    # Final safety check on content size
    combined_content = f"CONTROL: {control_desc}\n\nEVIDENCE: {evidence_content}"
    if len(combined_content) > MAX_TOTAL_CHARS:
        # Prioritize control description, truncate evidence if needed
        available_for_evidence = MAX_TOTAL_CHARS - len(control_desc) - 100
        if available_for_evidence > 500:
            evidence_content = smart_truncate_content(evidence_content, available_for_evidence)
        else:
            evidence_content = evidence_content[:500] + "[TRUNCATED due to size limits]"
        print(f"  ⚠️ Evidence content truncated for sufficiency assessment")
    
    prompt = f"""
You are a world-class auditor performing a Test of Effectiveness (TOE) review. 

Given the control description below and the evidence provided, perform a thorough assessment:

1. Start with a clear YES or NO - Is the evidence sufficient to conclude that this control is operating effectively?

2. Provide detailed auditor-style reasoning including:
   - Specific evidence that supports control effectiveness
   - Any gaps or deficiencies identified
   - Whether the evidence demonstrates the control operated as designed
   - Adequacy of documentation, timing, authorization
   - Overall assessment of control operation

CONTROL DESCRIPTION:
{control_desc}

EVIDENCE PROVIDED:
{evidence_content}

Provide your assessment in the format:
CONCLUSION: [YES/NO]

DETAILED REASONING:
[Your comprehensive auditor assessment]
"""
    
    payload = {
        "model": Config.MODEL,
        "messages": [
            {"role": "system", "content": "You are an expert auditor performing Test of Effectiveness reviews. Be thorough and professional."},
            {"role": "user", "content": prompt}
        ],
        "max_tokens": 2048,
        "temperature": 0.7
    }
    
    return make_llm_request_with_retry(client, payload, "Evidence Sufficiency")

# ───── Load Excel Input File ───────────────────────────────────
try:
    df = pd.read_excel(Config.INPUT_FILE, engine="openpyxl")
    required_cols = ["Control", "Control Description"]
    for col in required_cols:
        if col not in df.columns:
            sys.exit(f"⚠ Missing required column: {col}")
except Exception as e:
    sys.exit(f"⚠ Could not read {Config.INPUT_FILE} → {e}")

print(f"✔ Loaded {len(df)} controls from {Config.INPUT_FILE}")

# Initialize OpenAI client
client = initialize_client()

# ───── Process Each Control ────────────────────────────────────
evidence_summary_col = []
evidence_sufficiency_col = []

print(f"\nProcessing {len(df)} controls...")
print(f"Token Management: Max {MAX_FILE_CHARS} chars/file, {MAX_TOTAL_CHARS} chars/control")
print("="*60)

total_estimated_tokens = 0

for idx, (_, row) in enumerate(df.iterrows()):
    control_name = row.get('Control', '')
    control_desc = row.get('Control Description', '')
    
    progress = f"[{idx+1}/{len(df)}]"
    print(f"{progress} Processing: {control_name}")
    
    # Read evidence from folder
    print(f"  📂 Reading evidence folder...")
    evidence_content = read_evidence_folder(control_name)
    
    if "No evidence folder found" in evidence_content:
        print(f"  ⚠️ {evidence_content}")
        summary = "No evidence folder found for this control"
        sufficiency = "CONCLUSION: NO\n\nDETAILED REASONING: No evidence was provided for review. Cannot assess control effectiveness without supporting documentation."
    else:
        # Generate evidence summary
        print("  🔍 Analyzing evidence content...")
        summary = ask_evidence_summary(client, evidence_content)
        
        # Assess evidence sufficiency
        print("  ⚖️ Assessing evidence sufficiency...")
        sufficiency = ask_evidence_sufficiency(client, control_desc, evidence_content)
        
        # Track token usage
        control_tokens = estimate_tokens(evidence_content + control_desc)
        total_estimated_tokens += control_tokens
        print(f"  ✅ Analysis complete (~{control_tokens} tokens)")
    
    evidence_summary_col.append(summary)
    evidence_sufficiency_col.append(sufficiency)
    print()  # Add blank line between controls

print("="*60)
print(f"🏁 Processing completed!")
print(f"📊 Total estimated tokens used: ~{total_estimated_tokens}")
print(f"💰 Approximate cost estimation: ${total_estimated_tokens * 0.000015:.4f} (at $15/1M tokens)")
print("="*60)

# ───── Add Results to DataFrame ────────────────────────────────
df['Evidence Content Analysis'] = evidence_summary_col
df['Evidence Sufficiency Assessment'] = evidence_sufficiency_col

# ───── Save Output Excel File ──────────────────────────────────
print("Saving results...")
with pd.ExcelWriter(Config.OUTPUT_FILE, engine="openpyxl") as xl:
    df.to_excel(xl, sheet_name="TOE Evidence Analysis", index=False, startrow=2)
    wb = xl.book
    ws = wb["TOE Evidence Analysis"]
    
    # Insert header rows
    ws.insert_rows(1, amount=2)
    
    # Define input and output columns
    input_cols = ["Control", "Control Description"]
    output_cols = ["Evidence Content Analysis", "Evidence Sufficiency Assessment"]
    
    # Add section headers
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(input_cols))
    ws.cell(row=1, column=1).value = "INPUT COLUMNS"
    ws.merge_cells(start_row=1, start_column=len(input_cols)+1, end_row=1, end_column=len(input_cols)+len(output_cols))
    ws.cell(row=1, column=len(input_cols)+1).value = "OUTPUT COLUMNS"
    
    # Format all cells
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align
    
    # Auto-adjust column widths
    for col_idx, _ in enumerate(ws[3], 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)
    
    wb.save(Config.OUTPUT_FILE)

print(f"✅ TOE Evidence Analysis Results Saved: {Config.OUTPUT_FILE}")
print(f"\n" + "="*60)
print(f"🚀 AUTOMATED AUDITING TOOL - FEATURE SUMMARY:")
print(f"="*60)
print(f"✓ Multi-source evidence integration (Local files + SAP GRC + Jira)")
print(f"✓ Smart token management: {MAX_FILE_CHARS} chars/file, {MAX_TOTAL_CHARS} total")
print(f"✓ Intelligent content truncation preserving key information")
print(f"✓ Advanced error handling with exponential backoff")
print(f"✓ File size prioritization (smaller files processed first)")
print(f"✓ Enhanced processing for large Excel files and PDFs")
print(f"✓ OCR optimization for images and scanned documents")
print(f"✓ Email content processing (MSG, EML, MBOX formats)")
print(f"✓ Real-time token estimation and cost tracking")
print(f"✓ Robust retry logic for API rate limits")
print(f"✓ SAP GRC system integration for control data")
print(f"✓ Jira ticket integration for audit trail evidence")
print(f"\n" + "="*60)
print(f"📁 SETUP INSTRUCTIONS:")
print(f"="*60)
print(f"1. CONFIGURE YOUR SETTINGS:")
print(f"   • Set OPENAI_API_KEY environment variable or update API_KEY in script")
print(f"   • Update EVIDENCE_ROOT path: Current = '{Config.EVIDENCE_ROOT}'")
print(f"   • Update INPUT_FILE path: Current = '{Config.INPUT_FILE}'")
print(f"\n2. ENABLE SAP GRC INTEGRATION (Optional):")
print(f"   • Set SAP_GRC_ENABLED = True")
print(f"   • Configure SAP_GRC_URL, SAP_GRC_USERNAME, SAP_GRC_PASSWORD")
print(f"   • Set environment variables: SAP_GRC_USER, SAP_GRC_PASS")
print(f"\n3. ENABLE JIRA INTEGRATION (Optional):")
print(f"   • Set JIRA_ENABLED = True")
print(f"   • Configure JIRA_URL, JIRA_PROJECT_KEY")
print(f"   • Set environment variables: JIRA_USER, JIRA_TOKEN")
print(f"   • Create Jira API token at: https://id.atlassian.com/manage-profile/security/api-tokens")
print(f"\n4. ORGANIZE EVIDENCE FOLDERS:")
print(f"   Expected folder structure:")
print(f"     {Config.EVIDENCE_ROOT}/")
print(f"       ├── Control_Name_1/")
print(f"       │   ├── evidence_file1.pdf")
print(f"       │   ├── evidence_file2.docx")
print(f"       │   ├── email_evidence.msg")
print(f"       │   └── evidence_file3.xlsx")
print(f"       ├── Control_Name_2/")
print(f"       │   └── evidence_files...")
print(f"       └── ...")
print(f"\n5. SUPPORTED FILE FORMATS:")
print(f"   • Text files (TXT)")
print(f"   • Microsoft Office (DOCX, XLSX, XLS)")
print(f"   • PDF documents (with OCR fallback)")
print(f"   • CSV data files")
print(f"   • Email formats (MSG, EML, MBOX)")
print(f"   • Image files (PNG, JPG, etc. - with OCR)")
print(f"\n6. INSTALL REQUIRED PACKAGES:")
print(f"   pip install pandas openpyxl openai httpx python-docx PyPDF2")
print(f"   pip install extract-msg pytesseract pillow pdf2image requests")
print(f"\n7. INTEGRATION BENEFITS:")
print(f"   • SAP GRC: Automatic control data and test results retrieval")
print(f"   • Jira: Audit trail tickets and issue tracking integration")
print(f"   • Combined analysis provides comprehensive evidence assessment")
print(f"\n8. The script automatically matches folder names to control names")
print(f"   and provides detailed AI-powered evidence analysis!")
print(f"="*60)

# ═══════════════════════════════════════════════════════════════════════════════
#                               DATA PROCESSING
# ═══════════════════════════════════════════════════════════════════════════════

def load_and_validate_data() -> pd.DataFrame:
    """Load Excel file and validate required columns"""
    try:
        df = pd.read_excel(Config.INPUT_FILE, engine="openpyxl")
        required_cols = ["Control", "Control Description"]
        for col in required_cols:
            if col not in df.columns:
                logger.error(f"Missing required column: {col}")
                sys.exit(1)
        logger.info(f"✔ Loaded {len(df)} controls from {Config.INPUT_FILE}")
        return df
    except Exception as e:
        logger.error(f"Could not read {Config.INPUT_FILE}: {e}")
        sys.exit(1)

def process_controls(client: OpenAI, df: pd.DataFrame) -> pd.DataFrame:
    """Process all controls and generate analysis results"""
    summary_list = []
    sufficiency_list = []
    
    logger.info(f"Starting analysis of {len(df)} controls...")
    start_time = time.time()
    
    for idx, row in df.iterrows():
        control_name = str(row.get("Control", f"Control_{idx}"))
        control_desc = str(row.get("Control Description", "No description provided"))
        
        logger.info(f"  🔍 Processing: {control_name}")
        
        try:
            # Gather evidence
            evidence_content = read_evidence_folder(control_name)
            
            # Analyze evidence
            summary = ask_evidence_summary(client, evidence_content)
            sufficiency = ask_evidence_sufficiency(client, control_desc, evidence_content)
            
            summary_list.append(summary)
            sufficiency_list.append(sufficiency)
            
            logger.info(f"  ✓ Completed analysis for {control_name}")
            
        except Exception as e:
            logger.error(f"  ✗ Error processing {control_name}: {e}")
            summary_list.append(f"Error during analysis: {e}")
            sufficiency_list.append(f"Error during analysis: {e}")
    
    # Add results to DataFrame
    df["Evidence Summary"] = summary_list
    df["Evidence Sufficiency Assessment"] = sufficiency_list
    
    elapsed_time = time.time() - start_time
    logger.info(f"✓ Analysis completed in {elapsed_time:.1f} seconds")
    return df

def save_results_to_excel(df: pd.DataFrame):
    """Save results to Excel with formatting"""
    logger.info(f"Saving results to {Config.OUTPUT_FILE}")
    
    try:
        with pd.ExcelWriter(Config.OUTPUT_FILE, engine="openpyxl") as xl:
            df.to_excel(xl, sheet_name="TOE Results", index=False, startrow=2)

            wb = xl.book
            ws = wb["TOE Results"]

            # Insert two rows at the top
            ws.insert_rows(1, amount=2)
            
            # Identify input and output columns
            input_cols = ["Risk", "Risk Description", "Control", "Control Description"]
            output_cols = [col for col in df.columns if col not in input_cols]
            
            # Merge and label columns
            if len(input_cols) > 0:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(input_cols))
                ws["A1"] = "INPUT COLUMNS"
                ws["A1"].font = Font(bold=True)
                ws["A1"].alignment = Alignment(horizontal="center")
            
            if len(output_cols) > 0:
                output_start_col = len(input_cols) + 1
                ws.merge_cells(start_row=1, start_column=output_start_col, end_row=1, end_column=output_start_col + len(output_cols) - 1)
                # Set value before merging
                ws.cell(row=1, column=output_start_col, value="OUTPUT COLUMNS")
                output_cell = ws.cell(row=1, column=output_start_col)
                output_cell.font = Font(bold=True)
                output_cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for col_num, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"✓ Results saved to: {Config.OUTPUT_FILE}")
    except Exception as e:
        logger.error(f"Error saving results: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
#                               MAIN EXECUTION
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    """Main execution function"""
    start_time = time.time()
    
    try:
        logger.info("="*80)
        logger.info("AI-Powered TOE Evidence Analysis Framework")
        logger.info("="*80)
        
        # Initialize client
        client = initialize_client()
        
        # Load and validate data
        df = load_and_validate_data()
        
        # Process controls
        df_results = process_controls(client, df)
        
        # Save results
        save_results_to_excel(df_results)
        
        # Summary
        elapsed_time = time.time() - start_time
        logger.info("="*80)
        logger.info(f"✓ Analysis completed successfully!")
        logger.info(f"✓ Processed {len(df)} controls in {elapsed_time:.1f} seconds")
        logger.info(f"✓ Results saved to: {Config.OUTPUT_FILE}")
        logger.info("="*80)
        
    except KeyboardInterrupt:
        logger.info("\n⚠ Analysis interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"✗ Analysis failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()